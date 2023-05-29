import datetime
import tkinter
from tkinter import ttk
from tkinter import *
from tkinter import messagebox
import tkinter.font
import os
import openpyxl
import subprocess, os, platform
import time
import json
import re


# if "yigit Akoymak" in User_Names:
#    filepath = r"C:\Users\yigit\Desktop\excel_data\kk.xlsx"
#    workbook = openpyxl.load_workbook(filepath)
#    sheet = workbook.active
#
#    workbook.save(filepath)
class DataEntry:
    def __init__(self):
        self.window = tkinter.Tk()
        self.window.configure(background='#eff0f1')
        # install a new theme called awdark
        self.window.tk.call('source', 'tkBreeze-master/breeze/breeze.tcl')
        self.window.state('zoomed')
        # apply the theme
        self.s = ttk.Style()
        self.s.theme_use('breeze')
        self.s.configure("my.TButton", font=("Hack", 14))
        # title of the window
        self.window.title("Data Entry Sepkon")
        self.frame = ttk.Frame(self.window)
        # fonts
        ttk_font = tkinter.font.Font(size=10, font="Hack")
        ttk_font_header = tkinter.font.Font(size=28, font="Hack")
        ttk_entry_font = tkinter.font.Font(size=8, font="Hack")
        # Buyer LabelFrame
        self.buyer_font = ttk.Label(text="Buyer", font=24)
        self.buyer = ttk.LabelFrame(self.window, labelwidget=self.buyer_font)
        self.buyer.grid(row=0, column=0, sticky="NW")
        # Buyer info

        self.buyer_info_label = ttk.Label(self.buyer, text='Buyer name', font=ttk_font)
        self.buyer_info_combobox = ttk.Combobox(self.buyer)
        self.buyer_info_label.grid(row=0, column=1)
        self.buyer_info_combobox.grid(row=0, column=2)

        self.buyer_info_combobox.bind("<KeyRelease>", self._check_key)
        # self.buyer_info =
        # Saving User Info
        self.header_font = ttk.Label(text="Header", font=24)
        self.user_info_frame = ttk.LabelFrame(self.window, labelwidget=self.header_font)
        self.user_info_frame.grid(row=1, column=0, sticky='NW')

        self.Open_Data_label = ttk.Label(self.user_info_frame, text="Open Date", font=ttk_font)
        self.Open_Data_label.grid(row=0, column=0)
        self.Dead_Line_label = ttk.Label(self.user_info_frame, text="Dead Line", font=ttk_font)
        self.Dead_Line_label.grid(row=0, column=1)

        self.Open_Data_name_entry = ttk.Entry(self.user_info_frame, font=ttk_entry_font)
        self.Dead_Line_entry = ttk.Entry(self.user_info_frame, font=ttk_entry_font)
        self.Open_Data_name_entry.grid(row=1, column=0)
        self.Dead_Line_entry.grid(row=1, column=1)

        self.Request_Type_label = ttk.Label(self.user_info_frame, text="Request Type", font=ttk_font)
        self.Request_Type_Entry = ttk.Entry(self.user_info_frame, font=ttk_entry_font)
        self.Request_Type_label.grid(row=0, column=2)
        self.Request_Type_Entry.grid(row=1, column=2)
        # tax Exception section
        self.Tax_Exception_Label = ttk.Label(self.user_info_frame, text="Tax Exception", font=ttk_font)
        self.Tax_Exception_Entry = ttk.Entry(self.user_info_frame, font=ttk_entry_font)
        self.Tax_Exception_Label.grid(row=2, column=0)
        self.Tax_Exception_Entry.grid(row=3, column=0)

        # Required Delivery
        self.Required_Delivery_Label = ttk.Label(self.user_info_frame, text="Requited Delivery", font=ttk_font)
        self.Required_Delivery = ttk.Entry(self.user_info_frame, font=ttk_entry_font)
        self.Required_Delivery_Label.grid(row=2, column=1)
        self.Required_Delivery.grid(row=3, column=1)
        # Origin Restriction
        self.Origin_Restiriction = ttk.Label(self.user_info_frame, text="Origin Restriction", font=ttk_font)
        self.Origin_Restiriction_entry = ttk.Entry(self.user_info_frame, font=ttk_entry_font)
        self.Origin_Restiriction.grid(row=2, column=2)
        self.Origin_Restiriction_entry.grid(row=3, column=2)
        # Operation Restriction
        self.Operation_Restiriction = ttk.Label(self.user_info_frame, text="Operation Restriction", font=ttk_font)
        self.Operation_Restiriction_entry = ttk.Entry(self.user_info_frame, font=ttk_entry_font)
        self.Operation_Restiriction.grid(row=4, column=0)
        self.Operation_Restiriction_entry.grid(row=5, column=0)
        # Operation Type
        self.Operation_Type = ttk.Label(self.user_info_frame, text="Operation Type", font=ttk_font)
        self.Operation_Type_entry = ttk.Entry(self.user_info_frame, font=ttk_entry_font)
        self.Operation_Type.grid(row=4, column=1)
        self.Operation_Type_entry.grid(row=5, column=1)
        # Project end use
        self.Project_End_Use_Label = ttk.Label(self.user_info_frame, text="Project end use", font=ttk_font)
        self.Project_End_Use_Entry = ttk.Entry(self.user_info_frame, font=ttk_entry_font)
        self.Project_End_Use_Label.grid(row=4, column=2)
        self.Project_End_Use_Entry.grid(row=5, column=2)

        for widget in self.user_info_frame.winfo_children():
            widget.grid_configure(padx=5, pady=5)

        for widget in self.buyer.winfo_children():
            widget.grid_configure(padx=5, pady=5)

        # Saving Course Info
        self.courses_frame_font = ttk.Label(text="Main", font=ttk_font_header)
        self.courses_frame = ttk.LabelFrame(self.window, labelwidget=self.courses_frame_font)
        self.courses_frame.grid(row=2, column=0, sticky='w')

        # Postion of the excel values
        self.pos_label = ttk.Label(self.courses_frame, text="POS", font=ttk_font)
        self.pos_entry = ttk.Entry(self.courses_frame, font=ttk_font)
        self.pos_label.grid(row=0, column=0)
        self.pos_entry.grid(row=1, column=0)

        # ERP code
        self.ERP_Grup_label = ttk.Label(self.courses_frame, text="ERP_Grup", font=ttk_font)
        self.ERP_Entry = ttk.Entry(self.courses_frame, font=ttk_font)
        self.ERP_Grup_label.grid(row=0, column=1)
        self.ERP_Entry.grid(row=1, column=1)

        # Description
        self.Description_label = ttk.Label(self.courses_frame, text="Description", font=ttk_font)
        self.Description_box = ttk.Entry(self.courses_frame, font=ttk_font)
        self.Description_label.grid(row=2, column=0)
        self.Description_box.grid(row=3, column=0, sticky='news')

        # Dimensions_1
        self.Dimensions_label = ttk.Label(self.courses_frame, text="Dim_1", font=ttk_font)
        self.Dimensions_box = ttk.Entry(self.courses_frame, font=ttk_font)
        self.Dimensions_label.grid(row=4, column=0)
        self.Dimensions_box.grid(row=5, column=0)

        # Dimension 2
        self.Dimensions2_label = ttk.Label(self.courses_frame, text="Dim_2", font=ttk_font)
        self.Dimensions2_box = ttk.Entry(self.courses_frame, font=ttk_font)
        self.Dimensions2_label.grid(row=4, column=1)
        self.Dimensions2_box.grid(row=5, column=1)

        # Dimension 3
        self.Dimensions3_label = ttk.Label(self.courses_frame, text="Dim_3", font=ttk_font)
        self.Dimensions3_box = ttk.Entry(self.courses_frame, font=ttk_font)
        self.Dimensions3_label.grid(row=6, column=0)
        self.Dimensions3_box.grid(row=7, column=0)

        # L-mm values
        self.l_mm_label = ttk.Label(self.courses_frame, text="L-mm", font=ttk_font)
        self.l_mm_box = ttk.Entry(self.courses_frame, font=ttk_font)
        self.l_mm_label.grid(row=6, column=1)
        self.l_mm_box.grid(row=7, column=1)

        # Quantity
        # self.kg_p_label = ttk.Label(self.courses_frame,t)
        for widget in self.courses_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Accept terms
        self.terms_frame_font = ttk.Label(text='Terms & Conditions', font=24)
        self.terms_frame = ttk.LabelFrame(self.window, labelwidget=self.terms_frame_font)
        self.terms_frame.grid(row=3, column=0, sticky='nwes')
        self.accept_var_font = ttk.Label(text="I accept the terms and conditions.", font=14)
        self.accept_var = tkinter.StringVar(value="Not Accepted")
        self.terms_check = ttk.Checkbutton(self.terms_frame, text="I accept the terms and conditions.",
                                           variable=self.accept_var, onvalue="Accepted", offvalue="Not Accepted")
        self.terms_check.grid(row=0, column=0)

        # Button enter_data
        self.button = ttk.Button(self.window, text="enter data", style='my.TButton', command=self.enter_data)
        self.button.grid(row=4, column=0, sticky="news", padx=20, pady=10)

        self.window.mainloop()

    # -Buyer_combobox-#
    def _from_json_to_dict(self):
        with open('user_names.json', 'r') as file:
            json_data = file.read()
            data_user = json.loads(json_data)
        return data_user

    def _buyers_list(self) -> list:
        # adress F11, tellephone F13,web F14, contact person F16, Email F17
        usr_sheet_local = ['F11', 'F13', 'F14', 'F16', 'F17']
        user_keys = list(self._from_json_to_dict().keys())
        return user_keys

    def _insert_buyer_info_excel(self):
        pass

    def _update(self, data):
        self.buyer_info_combobox['values'] = data

    def _check_key(self, event):
        value = self.buyer_info_combobox.get()
        # get data from la
        if value == "":
            data = self._buyers_list()
        else:
            regex = fr'^{value}'
            data = [string for string in self._buyers_list() if re.match(regex, string)]
        # update data in combobox
        self._update(data)

    ###################################################################################################################
    def Date_time(self) -> list:
        """get the month and day of the current time"""
        date = datetime.datetime.now()
        return [str(date.day), str(date.month), str(date.year)[-2:]]

    def enter_data(self):
        accepted = self.accept_var.get()
        if accepted == "Accepted":
            # Get the header values
            open_date = self.Open_Data_name_entry.get()
            dead_line = self.Dead_Line_entry.get()
            request_type = self.Request_Type_Entry.get()
            tax_exception = self.Tax_Exception_Entry.get()
            required_delivery = self.Required_Delivery.get()
            origin_restiriction = self.Origin_Restiriction_entry.get()
            operation_Restriction = self.Operation_Restiriction_entry.get()
            operation_type = self.Operation_Type_entry.get()
            project_end_use = self.Project_End_Use_Entry.get()
            buyer_name = self.buyer_info_combobox.get()
            # Get the main values
            postion = self.pos_entry.get()
            erp = self.ERP_Entry.get()
            description = self.Description_box.get()
            dim1 = self.Dimensions_box.get()
            dim2 = self.Dimensions2_box.get()
            dim3 = self.Dimensions3_box.get()
            l_mm = self.l_mm_box.get()
            print(postion, erp, description, dim2, dim1, dim3)
            # Course info
            # registration_status = self.reg_status_var.get()
            # numcourses = self.pos_entry.get()
            # numsemesters = self.ERP_Entry.get()

            # print("First name: ", open_date, "ERP_GROUP: ", dead_line)
            # print("Title: ", request_type, "Age: ", tax_exception, "Nationality: ")
            # print("# Courses: ", numcourses, "# Semesters: ", numsemesters)
            # print("Registration status", registration_status)
            # print("------------------------------------------")

            filepath = r"C:\Users\yigit\Desktop\excel_data\kk.xlsx"

            if not os.path.exists(filepath):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                heading = ["First Name", tax_exception, "Erp_group", "Title", "Age", "Nationality",
                           "# Courses", "# Semesters", "Registration status"]
                sheet.append(heading)
                workbook.save(filepath)
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            # add buyer name to excel
            bynm = sheet['F9']
            bynm.value = buyer_name
            data = self._buyers_list()[f"{buyer_name}"]
            for i in range(len(data)):
                cell_location = usr_sheet_loca[i]
                cell = sheet[cell_location]
                cell.value = data[i][0]
            # Add Open date and deadline
            opdv, ddl = sheet['S10'], sheet['S11']
            opdv.value, ddl.value = open_date, dead_line
            # Request Type, Tax exceptation
            rtyp, txex = sheet['S12'], sheet['S13']
            rtyp.value, txex.value = request_type, tax_exception
            # Requred delivery, Origin Restiriction
            rqd, org = sheet['S14'], sheet['S15']
            rqd.value, org.value = required_delivery, origin_restiriction
            # Operation type, Origin Restriciton
            # opt = sheet['S16']
            # opr.value, opt.value = operation_Restriction, origin_restiriction
            # Add time to the excel
            time = sheet['S8']
            time.value = "/ ".join(self.Date_time())
            #
            sheet.append([open_date, dead_line, request_type, tax_exception])
            workbook.save(filepath)
            if platform.system() == 'Windows':
                os.startfile(filepath)
        else:
            tkinter.messagebox.showwarning(title="Error", message="You have not accepted the terms")


def PdfViewer(self):
    pass
    # root = tkinter.Tk()
    # create object like this.
    #
    # Add your pdf location and width and height.
    # variable2 = variable1.pdf_view(root, pdf_location=r"location", width=50, height=100)
    # variable2.pack()
    # root.mainloop()


DataEntry()

# User_Names = {"yigit Akoymak": [["London, Greenwich, SE10"],
#                                ["+90 530 068 89 48"],
#                                ["https://www.yigitakoymak.xyz"],
#                                ["Erdogan Akoymak"],
#                                ["akoymakyigit@gmail.com"]]
#
#              }
